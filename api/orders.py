import os
import pandas as pd
from io import BytesIO
from fastapi import APIRouter, Request, HTTPException
from database import db
from typing import List, Optional
from schemas import OrderUpdate, OrderNoteCreate
from utils import get_image_path

router = APIRouter()

@router.get("/orders")
def get_orders(q: Optional[str] = None, estatus: str = "", tipo: str = "", desde: str = "", hasta: str = "", mueble: str = ""):
    conn = db()
    cur = conn.cursor(dictionary=True)
    try:
        q_clean = (q or "").strip()
        estatus_clean = (estatus or "").strip()
        tipo_clean = (tipo or "").strip()

        sql = """
            SELECT
                o.id, o.folio, o.created_at, o.vendedor,
                o.total, o.anticipo_pagado, o.saldo,
                o.entrega_estimada, o.estatus, o.tipo, o.nota, o.estatus_solicitado,
                o.cp_envio, o.costo_envio, o.quote_id,
                q.cliente_nombre,
                (SELECT content FROM order_notes WHERE order_id = o.id ORDER BY id DESC LIMIT 1) as ultima_nota
            FROM orders o
            LEFT JOIN quotes q ON q.id = o.quote_id
            WHERE 1=1
        """
        params = []

        if q_clean:
            sql += """ AND (o.folio LIKE %s OR IFNULL(q.cliente_nombre,'') LIKE %s 
                       OR o.quote_id IN (SELECT quote_id FROM quote_lines ql JOIN products p ON p.id=ql.product_id WHERE p.modelo LIKE %s OR p.codigo LIKE %s))"""
            like = f"%{q_clean}%"
            params.extend([like, like, like, like])
        
        if estatus_clean:
            sql += " AND o.estatus = %s"
            params.append(estatus_clean)
            
        if tipo_clean:
            sql += " AND o.tipo = %s"
            params.append(tipo_clean)

        if desde:
            sql += " AND date(o.created_at) >= %s"
            params.append(desde)
        
        if hasta:
            sql += " AND date(o.created_at) <= %s"
            params.append(hasta)
            
        if mueble:
            sql += " AND o.quote_id IN (SELECT quote_id FROM quote_lines ql JOIN products p ON p.id=ql.product_id WHERE p.modelo LIKE %s OR p.codigo LIKE %s)"
            like_m = f"%{mueble}%"
            params.extend([like_m, like_m])

        sql += " ORDER BY o.created_at DESC LIMIT 200"
        
        cur.execute(sql, tuple(params))
        rows = cur.fetchall()
        return rows
    finally:
        conn.close()

@router.get("/orders/export")
def export_orders(format: str = "excel", q: Optional[str] = None, estatus: str = "", tipo: str = "", desde: str = "", hasta: str = "", mueble: str = ""):
    from fastapi.responses import StreamingResponse, Response
    import datetime

    # Fetch filtered data
    data = get_orders(q, estatus, tipo, desde, hasta, mueble)
    
    if not data:
        raise HTTPException(status_code=404, detail="No hay datos para exportar")

    if format == "pdf":
        from jinja2 import Environment, FileSystemLoader
        import os

        # Helper to decide badge classes
        def get_status_class(st):
            st = (st or "").upper()
            if st == "ENTREGADO": return "entregado"
            if st == "LIQUIDADO": return "liquidado"
            if st == "REGISTRADO": return "registrado"
            if st == "EN FABRICACIÓN": return "fabricacion"
            if st == "LISTO ENTREGA": return "listo"
            if st == "CANCELADO": return "cancelado"
            return "default"

        # Prep data for template
        for row in data:
            if isinstance(row["created_at"], str):
                # If it's a string like "2026-02-22T18:16:08", convert to "22/02/2026"
                try:
                    dt = datetime.datetime.fromisoformat(row["created_at"].replace('Z', '+00:00'))
                    row["fecha_str"] = dt.strftime("%d/%m/%Y")
                except:
                    row["fecha_str"] = row["created_at"].split()[0]
            else:
                row["fecha_str"] = row["created_at"].strftime("%d/%m/%Y")
            row["status_class"] = get_status_class(row["estatus"])

        filters_dict = {
            "q": q, "estatus": estatus, "mueble": mueble, "desde": desde, "hasta": hasta
        }
        has_filters = any(filters_dict.values())

        context = {
            "orders": data,
            "total_count": len(data),
            "sum_total": sum(r["total"] for r in data),
            "sum_saldo": sum(r["saldo"] for r in data),
            "filters": filters_dict,
            "has_filters": has_filters,
            "fecha_generacion": datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        }

        from utils import get_image_b64
        context["logo_b64"] = get_image_b64('logo.jpg') or get_image_b64('logo.png') or get_image_b64('logo.jpeg')
        context["fb_b64"] = get_image_b64('Facebook_logo.png')
        context["ig_b64"] = get_image_b64('Instagram_icon.png')
        context["wa_b64"] = get_image_b64('whatsapp_icon.png')

        env = Environment(loader=FileSystemLoader(os.path.join(os.path.dirname(__file__), 'templates')))
        template = env.get_template('orders_report.html')
        html_out = template.render(context)

        from weasyprint import HTML
        
        pdf = HTML(string=html_out).write_pdf()
        
        return Response(
            content=pdf,
            media_type="application/pdf",
            headers={"Content-Disposition": 'inline; filename="reporte_pedidos.pdf"'}
        )
    else:
        import pandas as pd
        from io import BytesIO
        import openpyxl
        import os
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
        from openpyxl.drawing.image import Image as OpenPyxlImage
        from utils import get_image_path
        from dateutil.relativedelta import relativedelta

        # 1. Fetch furniture items for all orders
        conn = db()
        cur = conn.cursor(dictionary=True)
        try:
            quote_ids = [r["quote_id"] for r in data if r.get("quote_id")]
            furniture_map = {}
            if quote_ids:
                format_strings = ','.join(['%s'] * len(quote_ids))
                cur.execute(f"""
                    SELECT ql.quote_id, p.modelo 
                    FROM quote_lines ql
                    JOIN products p ON p.id = ql.product_id
                    WHERE ql.quote_id IN ({format_strings})
                """, tuple(quote_ids))
                lines = cur.fetchall()
                for ln in lines:
                    qid = ln["quote_id"]
                    if qid not in furniture_map: furniture_map[qid] = []
                    furniture_map[qid].append(ln["modelo"])
        finally:
            conn.close()

        # 2. Process Data for Excel
        processed_data = []
        max_furnitures = 0
        for row in data:
            # Handle Dates
            created_at = row["created_at"]
            if isinstance(created_at, str):
                try:
                    created_at = datetime.datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                except:
                    # Fallback for other string formats
                    try:
                        created_at = datetime.datetime.strptime(created_at.split('.')[0], "%Y-%m-%d %H:%M:%S")
                    except:
                        created_at = datetime.datetime.now() # Fallback

            # Logic for Delivery Date
            tipo = row["tipo"]
            estatus = row["estatus"]
            
            # Fecha Promesa Calculation
            fecha_promesa_dt = None
            if tipo == "PEDIDO_FABRICACION":
                fecha_promesa_dt = created_at + datetime.timedelta(days=25)
            elif tipo == "APARTADO":
                fecha_promesa_dt = created_at + relativedelta(months=3)
            else: # VENTA_STOCK or others
                fecha_promesa_dt = created_at

            # If it's already "ENTREGADO", we show the "entrega_estimada" or the date it was delivered
            # For now, we'll prefix based on status
            if estatus == "ENTREGADO":
                entrega_info = row.get("entrega_estimada") or "Entregado"
            else:
                entrega_info = fecha_promesa_dt.strftime("%d/%m/%Y") if fecha_promesa_dt else (row.get("entrega_estimada") or "N/A")

            # Furniture items
            qid = row.get("quote_id")
            items = furniture_map.get(qid, [])
            max_furnitures = max(max_furnitures, len(items))

            new_row = {
                "Folio": row["folio"],
                "Fecha": created_at.replace(tzinfo=None) if isinstance(created_at, datetime.datetime) else created_at,
                "Vendedor": row["vendedor"],
                "Cliente": row["cliente_nombre"],
                "Tipo": row["tipo"],
                "Estado": row["estatus"],
                "Total": float(row["total"] or 0),
                "Anticipo": float(row["anticipo_pagado"] or 0),
                "Saldo": float(row["saldo"] or 0),
                "Fecha Entrega": entrega_info
            }

            # Add furniture columns
            for i, itm in enumerate(items):
                new_row[f"Mueble {i+1}"] = itm
            
            processed_data.append(new_row)

        df = pd.DataFrame(processed_data)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # We'll start the dataframe at row 10 to leave space for header/logo
            df.to_excel(writer, index=False, sheet_name='Pedidos', startrow=9)
            
            workbook = writer.book
            worksheet = writer.sheets['Pedidos']
            
            # 3. Add Logo
            # Try several paths based on find_by_name results
            base_p = os.path.dirname(os.path.abspath(__file__))
            logo_candidates = [
                os.path.join(base_p, 'static', 'img', 'logo.png'),
                os.path.join(base_p, 'static', 'img', 'logo.jpg'),
                os.path.join(os.path.dirname(base_p), 'static', 'logo.jpg'),
                get_image_path('logo.png'),
                get_image_path('logo.jpg')
            ]
            
            logo_path = None
            for p in logo_candidates:
                if os.path.exists(p):
                    logo_path = p
                    break
            
            if logo_path:
                try:
                    img = OpenPyxlImage(logo_path)
                    img.width = 150
                    img.height = 70
                    worksheet.add_image(img, 'A1')
                except:
                    pass

            # 4. Add Report Title and Filters
            worksheet['D2'] = "REPORTE DE PEDIDOS - LP MUEBLERÍA"
            worksheet['D2'].font = Font(size=16, bold=True, color="D4AF37")
            
            worksheet['D4'] = f"Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
            
            # Filters Info
            filter_text = []
            if q: filter_text.append(f"Búsqueda: {q}")
            if estatus: filter_text.append(f"Estado: {estatus}")
            if tipo: filter_text.append(f"Tipo: {tipo}")
            if desde: filter_text.append(f"Desde: {desde}")
            if hasta: filter_text.append(f"Hasta: {hasta}")
            if mueble: filter_text.append(f"Mueble: {mueble}")
            
            worksheet['D5'] = "Filtros aplicados: " + (", ".join(filter_text) if filter_text else "Ninguno")
            worksheet['D5'].font = Font(italic=True)

            # 5. Styling the Table
            header_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # Stylize headers (row 10)
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=10, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            # Stylize data rows and format currency
            for row_idx in range(11, 11 + len(df)):
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.border = border
                    
                    # Currency Formatting for Total, Anticipo, Saldo (Cols 7, 8, 9)
                    if col_idx in [7, 8, 9]:
                        cell.number_format = '"$" #,##0.00'
                        cell.alignment = Alignment(horizontal="right")

            # 6. Auto-adjust columns width
            for i, col in enumerate(df.columns):
                max_len = max(
                    df[col].astype(str).map(len).max() if not df.empty else 0,
                    len(str(col))
                ) + 3
                column_letter = get_column_letter(i + 1)
                worksheet.column_dimensions[column_letter].width = min(max_len, 50) # Cap at 50

        output.seek(0)
        filename = f"reporte_pedidos_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        headers = {'Content-Disposition': f'attachment; filename="{filename}"'}
        return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@router.get("/orders/{order_id}")
def get_order_detail(order_id: int):
    conn = db()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT o.*, q.cliente_nombre, q.cliente_tel
            FROM orders o
            LEFT JOIN quotes q ON q.id = o.quote_id
            WHERE o.id=%s
        """, (order_id,))
        o = cur.fetchone()
        if not o:
            raise HTTPException(status_code=404, detail="Pedido no encontrado")
            
        cur.execute("SELECT * FROM payments WHERE order_id=%s ORDER BY id DESC", (order_id,))
        payments = cur.fetchall()

        # Fetch lines from quote
        cur.execute("""
            SELECT ql.*, p.codigo, p.modelo, p.tamano
            FROM quote_lines ql JOIN products p ON p.id=ql.product_id
            WHERE ql.quote_id=%s
        """, (o["quote_id"],))
        lines = cur.fetchall()

        return {"order": o, "payments": payments, "lines": lines}
    finally:
        conn.close()

@router.patch("/orders/{order_id}")
async def update_order(order_id: int, data: OrderUpdate, request: Request):
    # Retrieve role from request - we'll assume it's passed or handled via headers/auth in actual implementation
    # For now, we'll check the 'X-Role' header if present, or default to admin for simplicity in tests
    role = request.headers.get("X-Role", "vendedor") 
    
    conn = db()
    cur = conn.cursor(dictionary=True)
    
    cur.execute("SELECT folio, vendedor FROM orders WHERE id=%s", (order_id,))
    o_info = cur.fetchone()
    if not o_info:
        conn.close()
        raise HTTPException(status_code=404, detail="Pedido no encontrado")
        
    updates = []
    params = []
    
    if data.estatus is not None:
        if data.estatus.upper() == "ENTREGADO":
            cur.execute("SELECT saldo FROM orders WHERE id=%s", (order_id,))
            oCheck = cur.fetchone()
            if oCheck and round(float(oCheck["saldo"]), 2) > 0:
                conn.close()
                raise HTTPException(status_code=400, detail="El pedido no está liquidado, no se puede cambiar a ENTREGADO")
                
        if role == "admin":
            updates.append("estatus = %s")
            params.append(data.estatus)
            updates.append("estatus_solicitado = ''") # Clear any pending request
        else:
            # Vendedor can only "solicit"
            updates.append("estatus_solicitado = %s")
            params.append(data.estatus)
            
            from api.notifications import trigger_notification
            trigger_notification(
                "admin", "warning", "Autorización Solicitada", 
                f"El vendedor {o_info['vendedor']} solicitó cambiar el pedido {o_info['folio']} a {data.estatus}", 
                order_id
            )
            
    if data.nota is not None:
        updates.append("nota = %s")
        params.append(data.nota)
        
    if data.entrega_estimada is not None:
        updates.append("entrega_estimada = %s")
        params.append(data.entrega_estimada)

    if data.entrega_promesa is not None:
        if role == "admin":
            updates.append("entrega_promesa = %s")
            params.append(data.entrega_promesa)
        # Vendedores cannot change entrega_promesa
    
    if hasattr(data, "cp_envio") and data.cp_envio is not None:
        updates.append("cp_envio = %s")
        params.append(data.cp_envio)
        
    if hasattr(data, "costo_envio") and data.costo_envio is not None:
        updates.append("costo_envio = %s")
        params.append(data.costo_envio)
        
    if not updates:
        conn.close()
        return {"message": "No changes"}
        
    sql = f"UPDATE orders SET {', '.join(updates)} WHERE id = %s"
    params.append(order_id)
    
    cur.execute(sql, tuple(params))
    conn.commit()
    conn.close()
    return {"status": "success"}

@router.post("/orders/{order_id}/authorize")
def authorize_order_status(order_id: int):
    conn = db()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT folio, estatus_solicitado FROM orders WHERE id=%s", (order_id,))
        o = cur.fetchone()
        if not o or not o["estatus_solicitado"]:
            raise HTTPException(status_code=400, detail="No hay cambio de estatus pendiente")
            
        new_status = o["estatus_solicitado"]
        cur.execute("UPDATE orders SET estatus=%s, estatus_solicitado='' WHERE id=%s", (new_status, order_id))
        
        from api.notifications import trigger_notification
        trigger_notification(
            "vendedor", "success", "Cambio Autorizado", 
            f"Se autorizó el cambio a {new_status} para el pedido {o['folio']}", 
            order_id
        )
        
        conn.commit()
        return {"status": "success", "new_status": new_status}
    finally:
        conn.close()

@router.get("/orders/{order_id}/notes")
def get_order_notes(order_id: int):
    conn = db()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT * FROM order_notes WHERE order_id=%s ORDER BY id DESC", (order_id,))
        notes = cur.fetchall()
        return notes
    finally:
        conn.close()

@router.post("/orders/{order_id}/notes")
def add_order_note(order_id: int, data: OrderNoteCreate):
    from utils import today_iso
    conn = db()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            INSERT INTO order_notes(order_id, author, content, created_at)
            VALUES (%s, %s, %s, %s)
        """, (order_id, data.author, data.content, today_iso()))
        conn.commit()
        return {"status": "success"}
    finally:
        conn.close()

@router.post("/orders/{order_id}/penalty")
async def apply_penalty(order_id: int, request: Request):
    role = request.headers.get("X-Role", "vendedor")
    conn = db()
    cur = conn.cursor(dictionary=True)
    
    try:
        cur.execute("SELECT total, saldo, folio, tipo FROM orders WHERE id=%s", (order_id,))
        o = cur.fetchone()
        if not o:
            raise HTTPException(status_code=404, detail="Pedido no encontrado")
        
        if o["tipo"] != "APARTADO":
            raise HTTPException(status_code=400, detail="Solo se pueden aplicar multas de quincena a pedidos de APARTADO.")

        penalty_amount = 200.0
        new_total = float(o["total"]) + penalty_amount
        new_saldo = float(o["saldo"]) + penalty_amount

        cur.execute("UPDATE orders SET total=%s, saldo=%s WHERE id=%s", (new_total, new_saldo, order_id))
        
        # Log this action
        author = "Admin" if role == "admin" else "Vendedor"
        cur.execute("INSERT INTO order_notes(order_id, author, content) VALUES (%s,%s,%s)", 
                    (order_id, author, f"🔴 SE APLICÓ MULTA POR ATRASO EN QUINCENA: +$200.00 MXN. (Total anterior: ${o['total']}, Nuevo: ${new_total})"))
        
        conn.commit()
        return {"msg": "Multa aplicada exitosamente"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()
