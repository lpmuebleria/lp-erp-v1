from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
import io
import os
import datetime
import pandas as pd
from database import db
from utils import compute_bolsas, get_image_path
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.drawing.image import Image as OpenPyxlImage
from openpyxl.utils import get_column_letter

def apply_premium_style(ws, title, filters_str, df_cols):
    """Utility to apply LP Mueblería brand styling to an openpyxl worksheet."""
    # 1. Add Logo
    base_p = os.path.dirname(os.path.abspath(__file__))
    logo_candidates = [
        os.path.join(base_p, 'static', 'img', 'logo.png'),
        os.path.join(base_p, 'static', 'img', 'logo.jpg'),
        os.path.join(os.path.dirname(base_p), 'static', 'logo.jpg'),
        get_image_path('logo.png'),
        get_image_path('logo.jpg')
    ]
    logo_path = next((p for p in logo_candidates if os.path.exists(p)), None)
    if logo_path:
        try:
            img = OpenPyxlImage(logo_path)
            img.width, img.height = 150, 70
            ws.add_image(img, 'A1')
        except: pass

    # 2. Title and Filters
    ws['D2'] = title
    ws['D2'].font = Font(size=16, bold=True, color="D4AF37")
    ws['D4'] = f"Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['D5'] = f"Filtros: {filters_str}"
    ws['D5'].font = Font(italic=True)

    # 3. Table Styling
    header_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Headers (row 10)
    for col_idx in range(1, len(df_cols) + 1):
        cell = ws.cell(row=10, column=col_idx)
        cell.fill, cell.font, cell.border = header_fill, header_font, border
        cell.alignment = Alignment(horizontal="center")

    # Data rows and Currency formating
    # We scan more rows than the initial DF because of potential footers added manually
    for row_idx in range(11, 2000): 
        # Check if row has data (by the first column which usually has Folio/Concepto)
        content = ws.cell(row=row_idx, column=1).value
        # If no content and we are far enough, stop
        if not content and row_idx > 11 + 20: break 
        if not content: continue

        for col_idx in range(1, len(df_cols) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            # Format if numeric and relevant column
            header_val = ws.cell(row=10, column=col_idx).value
            if isinstance(cell.value, (int, float)) and header_val not in ["Folio", "Fecha", "Cliente", "Estatus", "Descripción", "Concepto", "ID"]:
                cell.number_format = '"$"#,##0.00'

    # 4. Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value: max_length = max(max_length, len(str(cell.value)))
            except: pass
        ws.column_dimensions[column].width = min(max_length + 5, 50) # Cap at 50

router = APIRouter()

class ReportRequest(BaseModel):
    start_date: str
    end_date: str
    selected_families: List[str]
    include_expenses: bool = True

@router.post("/reports/team-activity")
def generate_team_activity_report(request: ReportRequest):
    conn = db()
    cur = conn.cursor(dictionary=True)

    try:
        # 1. Fetch Orders in date range
        cur.execute("""
            SELECT o.id, o.folio, o.created_at, o.estatus, o.anticipo_pagado, o.iva, q.cliente_nombre, q.id as quote_id, q.costo_envio
            FROM orders o
            JOIN quotes q ON o.quote_id = q.id
            WHERE DATE(o.created_at) BETWEEN %s AND %s
            AND o.estatus != 'CANCELADO'
            ORDER BY o.created_at DESC
        """, (request.start_date, request.end_date))
        orders = cur.fetchall()

        if not orders:
            # Return empty response early
            df_activity = pd.DataFrame(columns=["Folio", "Fecha", "Cliente", "Estatus"] + [f.capitalize() for f in request.selected_families])
            df_expenses = None
        else:
            # Batch fetch all lines for these quotes to avoid N+1 query
            quote_ids = tuple(o["quote_id"] for o in orders)
            format_quote_ids = ','.join(['%s'] * len(quote_ids))
            
            cur.execute(f"""
                SELECT ql.quote_id, ql.cantidad, ql.total_linea, p.tamano, p.costo_fabrica
                FROM quote_lines ql
                JOIN products p ON p.id = ql.product_id
                WHERE ql.quote_id IN ({format_quote_ids})
            """, quote_ids)
            all_lines = cur.fetchall()

            # Pre-fetch configs and settings
            cur.execute("SELECT k, v FROM settings WHERE k IN ('flete_unitario', 'comision_debito_pct', 'comision_msi_banco_pct', 'iva_automatico')")
            setts = {r["k"]: r["v"] for r in cur.fetchall()}
            flete_unitario = float(setts.get("flete_unitario", 0.0))
            iva_auto = int(setts.get("iva_automatico", 1))

            cur.execute("SELECT tamano, maniobras, empaque, comision, garantias FROM cost_config")
            config_rows = cur.fetchall()
            configs_by_tamano = {r["tamano"]: r for r in config_rows}
            default_cfg = {"maniobras": 0.0, "empaque": 0.0, "comision": 0.0, "garantias": 0.0}

            # Map quote_id -> computed_bolsas
            bolsas_by_quote = {}
            for q_id in quote_ids:
                 bolsas_by_quote[q_id] = {
                     "maniobras": 0.0, "empaque": 0.0, "comision": 0.0, "garantias": 0.0, 
                     "costo": 0.0, "venta": 0.0, "muebles": 0.0, "fletes": 0.0, "envios": 0.0, 
                     "iva": 0.0, "comision_tarjeta": 0.0, "comision_msi": 0.0
                 }

            for row in all_lines:
                qid = row["quote_id"]
                b = bolsas_by_quote[qid]
                cfg = configs_by_tamano.get(row["tamano"], default_cfg)
                
                qty = float(row["cantidad"] or 0)
                b["maniobras"] += float(cfg["maniobras"] or 0) * qty
                b["empaque"] += float(cfg["empaque"] or 0) * qty
                b["comision"] += float(cfg["comision"] or 0) * qty
                b["garantias"] += float(cfg["garantias"] or 0) * qty
                b["muebles"] += float(row["costo_fabrica"] or 0) * qty
                b["fletes"] += flete_unitario * qty
                b["costo"] += float(row["costo_fabrica"] or 0) * qty
                b["venta"] += float(row["total_linea"] or 0)

            # Priority order for money allocation (Waterfall)
            PRIORITY_ORDER = [
                "comision_msi",
                "comision_tarjeta",
                "iva",
                "comision",
                "muebles",
                "empaque",
                "garantias",
                "maniobras",
                "fletes",
                "envios",
                "utilidad_bruta"
            ]

            # Build data for the Activity Sheet
            activity_data = []
            totals = {fam: 0.0 for fam in request.selected_families}

            for o in orders:
                qid = o["quote_id"]
                bolsas_teoricas = bolsas_by_quote[qid]
                
                # Add envios (costo_envio) from order level
                bolsas_teoricas["envios"] = float(o["costo_envio"] or 0.0)
                
                # IVA Calculation (16% of sale, extracted if automatic)
                if iva_auto:
                    bolsas_teoricas["iva"] = bolsas_teoricas["venta"] - (bolsas_teoricas["venta"] / 1.16)
                else:
                    bolsas_teoricas["iva"] = bolsas_teoricas["venta"] * 0.16
                    
                # Fetch bank commissions for this order and split them
                cur.execute("SELECT metodo, monto, comision_bancaria FROM payments WHERE order_id=%s AND anulado=0", (o["id"],))
                payments = cur.fetchall()
                
                # Check if it's an MSI order (to potentially classify commissions if metodo is ambiguous)
                cur.execute("SELECT COUNT(*) as c FROM quote_lines WHERE quote_id=%s AND tipo_precio='msi'", (qid,))
                is_msi_order = (cur.fetchone()["c"] > 0)

                for p in payments:
                    mn = p["metodo"].lower().replace("_", " ").strip() if p["metodo"] else ""
                    saved_comm = float(p["comision_bancaria"] or 0.0)
                    monto_abono = float(p["monto"])

                    is_msi = "meses sin intereses" in mn or mn == "msi"
                    is_any_card = is_msi or "tarjeta" in mn or mn in ["debito", "credito", "td", "tc"]

                    # All installment plans (MSI) go to MSI bag (they include terminal 3% + bank 18% = 21% total)
                    if is_msi:
                        bolsas_teoricas["comision_msi"] += saved_comm
                    # All normal cards (Debit/Credit) go to 'comision_tarjeta'
                    elif is_any_card:
                        bolsas_teoricas["comision_tarjeta"] += saved_comm
                    elif is_msi_order and mn == 'transferencia':
                        pass

                bolsas_teoricas["utilidad_bruta"] = (
                    bolsas_teoricas["venta"] - bolsas_teoricas["iva"] - bolsas_teoricas["comision_tarjeta"] - bolsas_teoricas["comision_msi"] - bolsas_teoricas["costo"] - 
                    (bolsas_teoricas["maniobras"] + bolsas_teoricas["empaque"] + bolsas_teoricas["comision"] + bolsas_teoricas["garantias"] + bolsas_teoricas["fletes"])
                )

                # Waterfall (Cascada) Distribution
                anticipo = float(o["anticipo_pagado"] or 0.0)
                fondos_disponibles = anticipo
                
                bolsas_reales = {k: 0.0 for k in bolsas_teoricas.keys()}

                for bolsa in PRIORITY_ORDER:
                    monto_teorico = bolsas_teoricas.get(bolsa, 0.0)
                    
                    if fondos_disponibles <= 0:
                        bolsas_reales[bolsa] = 0.0
                    else:
                        if fondos_disponibles >= monto_teorico:
                            bolsas_reales[bolsa] = monto_teorico
                            fondos_disponibles -= monto_teorico
                        else:
                            bolsas_reales[bolsa] = fondos_disponibles
                            fondos_disponibles = 0.0
                            
                # Special cases: "venta" and "costo" are reference columns, not distribution buckets, so keep them as is if needed, 
                # but since they aren't in PRIORITY_ORDER we'll just copy them over (or keep logic clean for user selected families)
                bolsas_reales["venta"] = bolsas_teoricas["venta"]
                bolsas_reales["costo"] = bolsas_teoricas["costo"]
                
                row_act = {
                    "Folio": o["folio"],
                    "Fecha": o["created_at"],
                    "Cliente": o["cliente_nombre"],
                    "Estatus": o["estatus"]
                }
                
                has_relevant_data = False
                for fam in request.selected_families:
                    val = bolsas_reales.get(fam, 0.0)
                    row_act[fam.capitalize()] = val
                    totals[fam] += val
                    if val > 0:
                        has_relevant_data = True
                
                if has_relevant_data:
                    activity_data.append(row_act)

            if activity_data:
                totals_row = {"Folio": "TOTALES", "Fecha": "", "Cliente": "", "Estatus": ""}
                for fam in request.selected_families:
                    totals_row[fam.capitalize()] = totals[fam]
                activity_data.append(totals_row)

            df_activity = pd.DataFrame(activity_data)

        # 2. Fetch Expenses if requested
        df_expenses = None
        if request.include_expenses:
            # Create a tuple of selected families for the IN clause
            # We map families to lower/capitalized as they are stored in DB.
            # Usually stored exactly as typed or lowercase. Let's use lower for safety, or exact match.
            family_tuple = tuple(request.selected_families)
            
            if family_tuple:
                 format_strings = ','.join(['%s'] * len(family_tuple))
                 cur.execute(f"""
                     SELECT concepto, monto, descripcion, fecha 
                     FROM expenses 
                     WHERE fecha BETWEEN %s AND %s
                     AND concepto IN ({format_strings})
                     ORDER BY fecha DESC
                 """, (request.start_date, request.end_date, *family_tuple))
                 expenses = cur.fetchall()
                 
                 if expenses:
                     df_expenses = pd.DataFrame(expenses)
                     # Rename columns for presentation
                     df_expenses.rename(columns={
                         "concepto": "Concepto",
                         "monto": "Monto",
                         "descripcion": "Descripción",
                         "fecha": "Fecha"
                     }, inplace=True)
                     
                     # Add total row
                     total_expenses = df_expenses["Monto"].sum()
                     df_expenses.loc[len(df_expenses)] = ["TOTAL", total_expenses, "", ""]

        # 3. Generate Excel file with Professional Formatting
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # We'll start the dataframes at row 10 to leave space for header/logo
            df_activity.to_excel(writer, sheet_name='Actividad de Equipo', index=False, startrow=9)
            
            # Formatting the main sheet
            ws_act = writer.sheets['Actividad de Equipo']
            fam_str = ", ".join([f.capitalize() for f in request.selected_families])
            apply_premium_style(
                ws_act, 
                "REPORTE DE ACTIVIDAD OPERATIVA", 
                f"Desde {request.start_date} hasta {request.end_date} | Familias: {fam_str}",
                df_activity.columns
            )

            if request.include_expenses and df_expenses is not None and not df_expenses.empty:
                df_expenses.to_excel(writer, sheet_name='Gastos', index=False, startrow=9)
                ws_exp = writer.sheets['Gastos']
                apply_premium_style(
                    ws_exp, 
                    "REPORTE DE EGRESOS (GASTOS)", 
                    f"Desde {request.start_date} hasta {request.end_date} | Familias: {fam_str}",
                    df_expenses.columns
                )

        output.seek(0)
        filename = f"reporte_actividad_{request.start_date}_al_{request.end_date}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.post("/reports/cash-flow")
def generate_cash_flow_report(request: ReportRequest): # Reusing the ReportRequest schema but ignoring selected_families and include_expenses
    conn = db()
    cur = conn.cursor(dictionary=True)

    try:
        # 1. Fetch Incomes (Ingresos) from payments table
        cur.execute("""
            SELECT p.id, p.created_at as fecha, p.metodo, p.monto, p.referencia, o.folio as relacion
            FROM payments p
            JOIN orders o ON p.order_id = o.id
            WHERE DATE(p.created_at) BETWEEN %s AND %s
            AND p.anulado = 0
            ORDER BY p.created_at DESC
        """, (request.start_date, request.end_date))
        ingresos_raw = cur.fetchall()

        # 2. Fetch Expenses (Egresos) from expenses table
        cur.execute("""
            SELECT id, fecha, metodo_pago as metodo, monto, descripcion as referencia, concepto as relacion
            FROM expenses
            WHERE fecha BETWEEN %s AND %s
            ORDER BY fecha DESC
        """, (request.start_date, request.end_date))
        egresos_raw = cur.fetchall()

        # Processing totals
        ingresos_banco = 0.0
        ingresos_efectivo = 0.0
        egresos_banco = 0.0
        egresos_efectivo = 0.0

        for ing in ingresos_raw:
            if ing["metodo"] == "efectivo":
                ingresos_efectivo += float(ing["monto"])
            else: # transferencia, tarjeta_credito, tarjeta_debito
                ingresos_banco += float(ing["monto"])

        for eg in egresos_raw:
            if eg["metodo"] == "efectivo":
                egresos_efectivo += float(eg["monto"])
            else:
                egresos_banco += float(eg["monto"])

        # Create Dataframes
        df_ingresos = pd.DataFrame(ingresos_raw) if ingresos_raw else pd.DataFrame(columns=["id", "fecha", "metodo", "monto", "referencia", "relacion"])
        df_egresos = pd.DataFrame(egresos_raw) if egresos_raw else pd.DataFrame(columns=["id", "fecha", "metodo", "monto", "referencia", "relacion"])

        # Rename cols
        col_map = {"fecha": "Fecha", "metodo": "Método de Pago", "monto": "Monto", "referencia": "Referencia / Descripción", "relacion": "Pedido / Concepto"}
        df_ingresos.rename(columns=col_map, inplace=True, errors='ignore')
        df_egresos.rename(columns=col_map, inplace=True, errors='ignore')

        if not df_ingresos.empty and "id" in df_ingresos.columns: df_ingresos.drop(columns=["id"], inplace=True)
        if not df_egresos.empty and "id" in df_egresos.columns: df_egresos.drop(columns=["id"], inplace=True)

        # Totals Row for Incomes
        if not df_ingresos.empty:
            df_ingresos.loc[len(df_ingresos)] = ["TOTAL INGRESOS", "", sum(float(i["monto"]) for i in ingresos_raw), "", ""]
            
        # Totals Row for Expenses
        if not df_egresos.empty:
            df_egresos.loc[len(df_egresos)] = ["TOTAL EGRESOS", "", sum(float(e["monto"]) for e in egresos_raw), "", ""]

        # 3. Create Summary Cuadre DataFrame
        cuadre_data = [
            {"Concepto": "Ingresos en Efectivo", "Total": ingresos_efectivo},
            {"Concepto": "Menos: Egresos en Efectivo", "Total": -egresos_efectivo},
            {"Concepto": "FONDO FÍSICO ESPERADO (CAJA)", "Total": ingresos_efectivo - egresos_efectivo},
            {"Concepto": "", "Total": ""},
            {"Concepto": "Ingresos en Banco (Transf/Tarjetas)", "Total": ingresos_banco},
            {"Concepto": "Menos: Egresos pagados en Banco", "Total": -egresos_banco},
            {"Concepto": "SALDO BANCARIO ESPERADO (CUENTAS)", "Total": ingresos_banco - egresos_banco},
            {"Concepto": "", "Total": ""},
            {"Concepto": "BALANCE GLOBAL TOTAL", "Total": (ingresos_efectivo + ingresos_banco) - (egresos_efectivo + egresos_banco)}
        ]
        df_cuadre = pd.DataFrame(cuadre_data)

        # 4. Excel Generation with Styling
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_cuadre.to_excel(writer, sheet_name='RESUMEN DE CUADRE', index=False, startrow=9)
            df_ingresos.to_excel(writer, sheet_name='Desglose Ingresos', index=False, startrow=9)
            df_egresos.to_excel(writer, sheet_name='Desglose Egresos', index=False, startrow=9)

            # Apply Styles
            apply_premium_style(writer.sheets['RESUMEN DE CUADRE'], "FLUJO DE CAJA - RESUMEN DE CUADRE", f"Desde {request.start_date} hasta {request.end_date}", df_cuadre.columns)
            apply_premium_style(writer.sheets['Desglose Ingresos'], "FLUJO DE CAJA - INGRESOS DETALLADOS", f"Desde {request.start_date} hasta {request.end_date}", df_ingresos.columns)
            apply_premium_style(writer.sheets['Desglose Egresos'], "FLUJO DE CAJA - EGRESOS DETALLADOS", f"Desde {request.start_date} hasta {request.end_date}", df_egresos.columns)

        output.seek(0)
        filename = f"flujo_caja_{request.start_date}_al_{request.end_date}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()
